from src.libs import *
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill
import json
import os
import random

class StatisticsCalendar(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        # Setează titlul și dimensiunea ferestrei
        self.parent = parent
        self.parent.title("Select Date Range for Statistics")
        self.parent.geometry("600x600")

        # Path-ul directorului de date
        self.report_dir = "data/outputs/daily_report"
        
        # Obține intervalul de date disponibil
        self.min_date, self.max_date = self.get_date_range()

        # Crează componentele pentru calendar și buton
        tk.Label(self, text="Select Start Date:").pack(pady=10)
        self.start_calendar = Calendar(self, date_pattern="yyyy-mm-dd", mindate=self.min_date, maxdate=self.max_date)
        self.start_calendar.pack()

        tk.Label(self, text="Select End Date:").pack(pady=10)
        self.end_calendar = Calendar(self, date_pattern="yyyy-mm-dd", mindate=self.min_date, maxdate=self.max_date)
        self.end_calendar.pack()

        # Butonul pentru a genera statistici
        self.generate_button = tk.Button(self, text="Generate Statistics", command=self.generate_statistics)
        self.generate_button.pack(pady=20)

    def get_date_range(self):
        """Obține limita inferioară și superioară pe baza folderelor din data/outputs/daily_report"""
        try:
            folder_dates = []
            for folder_name in os.listdir(self.report_dir):
                folder_path = os.path.join(self.report_dir, folder_name)
                if os.path.isdir(folder_path):
                    try:
                        folder_date = datetime.strptime(folder_name, "%Y-%m-%d").date()
                        folder_dates.append(folder_date)
                    except ValueError:
                        continue

            if not folder_dates:
                raise ValueError("No valid date folders found.")

            return min(folder_dates), max(folder_dates)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to determine date range: {e}")
            self.parent.destroy()

    def generate_statistics(self):
        """Generează statistica și salvează fișierul Excel"""
        try:
            start_date = self.start_calendar.get_date()
            end_date = self.end_calendar.get_date()
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()

            # Verifică dacă data de început este mai mică decât data de sfârșit
            if start_date_obj > end_date_obj:
                messagebox.showerror("Error", "Start date must be earlier than end date.")
                return

            # Colectează fișierele JSON relevante
            relevant_files = []
            for folder_name in os.listdir(self.report_dir):
                folder_date = datetime.strptime(folder_name, "%Y-%m-%d").date()
                if start_date_obj <= folder_date <= end_date_obj:
                    folder_path = os.path.join(self.report_dir, folder_name)
                    for file_name in os.listdir(folder_path):
                        # Verificăm dacă numele fișierului este "average_statistics.json"
                        if file_name == "average_statistics.json":
                            relevant_files.append(os.path.join(folder_path, file_name))
            if not relevant_files:
                messagebox.showinfo("No Data", "No JSON files found for the selected date range.")
                return

            # Procesează fișierele JSON și generează statistica
            report_path = self.process_json_files(relevant_files, start_date, end_date)
            messagebox.showinfo("Success", f"Report generated successfully at: {report_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate statistics: {e}")


    def process_json_files(self, json_files, start_date, end_date):
        """Procesează fișierele JSON, le combină într-un fișier generalizat și generează raportul Excel cu histograme pentru toate câmpurile."""
        
        # Combina toate fișierele JSON într-un singur fișier generalizat
        combined_data = {}

        for file_path in json_files:
            with open(file_path, "r") as f:
                data = json.load(f)
                
                # Parcurge fiecare dată din fișierul JSON și o adaugă în fișierul generalizat
                for date, stats in data.items():
                    if isinstance(stats, dict):  # Verificăm dacă stats este un dicționar
                        if date not in combined_data:
                            combined_data[date] = stats
                        else:
                            # Dacă data există deja, combină valorile
                            for field, value in stats.items():
                                if isinstance(value, (int, float)):
                                    combined_data[date][field] = combined_data[date].get(field, 0) + value
                                else:
                                    combined_data[date][field] = value
                    else:
                        # Dacă stats nu este un dicționar (e.g., float), ignorăm și continuăm
                        print(f"Ignorăm valoarea pentru data {date} deoarece nu este un dicționar valid.")
        
        # Crează fișierul JSON generalizat
        output_dir = f"data/outputs/{start_date}_{end_date}"
        os.makedirs(output_dir, exist_ok=True)
        general_stats_file = os.path.join(output_dir, "general_stats.json")
        
        with open(general_stats_file, "w") as f:
            json.dump(combined_data, f, indent=4)

        # Structură pentru date
        fields_data = {field: {} for field in next(iter(combined_data.values())).keys()}

        # Adună datele pentru statistici
        for date, stats in combined_data.items():
            for field, value in stats.items():
                fields_data[field][date] = value

        # Sortează datele pe zile
        sorted_dates = sorted(next(iter(fields_data.values())).keys())

        # Crează fișierul Excel
        report_file = os.path.join(output_dir, "general_statistic.xlsx")
        wb = Workbook()

        # Sheet pentru tabelul cu date
        ws_data = wb.active
        ws_data.title = "Data Statistics"
        
        # Adaugă antet
        ws_data.append(["Date"] + list(fields_data.keys()))
        
        # Adaugă datele în Excel
        for date in sorted_dates:
            row = [date] + [fields_data[field].get(date, 0) for field in fields_data.keys()]
            ws_data.append(row)

        # Adaugă opțiune de sortare pentru fiecare coloană
        ws_data.auto_filter.ref = ws_data.dimensions
        
        # Sheet pentru histograme
        ws_histograms = wb.create_sheet("Histograms")
        
        # Generează histograme și le adaugă în sheet-ul de histograme
        for idx, (field, data) in enumerate(fields_data.items()):
            values = [data[date] for date in sorted_dates]
            
            # Creează histogramă
            plt.figure(figsize=(10, 5))
            color = [random.choice(["blue", "green", "red", "purple", "orange", "cyan"]) for _ in values]
            plt.bar(sorted_dates, values, color=color, label=f"{field} (Avg)", alpha=0.7)
            plt.xlabel("Date")
            plt.ylabel(field)
            plt.title(f"{field} Histogram")
            plt.legend()

            # Salvează histograma
            histogram_path = os.path.join(output_dir, f"{field}_histogram.png")
            plt.savefig(histogram_path)
            plt.close()

            # Adaugă imaginea la sheet-ul de histograme
            img = ExcelImage(histogram_path)
            img.width = 600
            img.height = 400
            col_position = chr(65 + idx * 2)  # Calculează poziția coloanei pentru fiecare histogramă (începe de la 'A')
            ws_histograms.add_image(img, f"{col_position}1")

        # Salvează fișierul Excel
        wb.save(report_file)

        return report_file