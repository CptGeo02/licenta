import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import tempfile
import matplotlib.ticker as ticker
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import pandas as pd
import tempfile
import matplotlib.pyplot as plt
import json
from openpyxl import Workbook
from datetime import datetime
import numpy as np
import os

class JsonToExcel:
    def __init__(self, tables_json_file, people_json_file, excel_file):
        self.tables_json_file = tables_json_file
        self.people_json_file = people_json_file
        self.excel_file = excel_file
        self.tables_performance_list = []
        self.people_number_list = []
        self.idle_time_ratio_list = []
        self.cleaning_ratio_list = []
        self.cycles_per_hour_list = []
        self.utilization_rate_list = []

        self.people_number_avg = 0
        self.tables_performance_avg = 0
        self.people_number_avg = 0
        self.idle_time_ratio_avg = 0
        self.cleaning_ratio_avg = 0
        self.cycles_per_hour_avg = 0
        self.utilization_rate_avg = 0

    def load_json(self, json_file):
        """Încarcă datele din fișierul JSON."""
        with open(json_file, 'r') as file:
            return json.load(file)

    def save_to_excel(self):
        """Salvează datele și graficele în Excel."""
         # Încărcarea datelor din JSON
        tables_data = self.load_json(self.tables_json_file)
        people_data = self.load_json(self.people_json_file)

        # Crearea DataFrame-urilor pentru datele meselor și oamenilor
        df_tables = pd.DataFrame(tables_data['table_statuses'])
        df_people = pd.DataFrame(people_data['detections'])

        # Adăugăm intervalele de timp
        df_tables = self.add_time_slot(df_tables)
        # Crearea unui workbook nou cu openpyxl
        workbook = Workbook()

        # Scrierea datelor brute în sheet-ul 'Raw Data'
        sheet = workbook.active
        sheet.title = 'Raw Data'
        for r, row in df_tables.iterrows():
            sheet.append(row.values.tolist())

        # Calcularea mediilor și generarea histogramelor pentru medii
        avg_durations = self.calculate_avg_durations(df_tables)
        self.generate_average_histogram(avg_durations, workbook)

        # Generarea histogramei și adăugarea ei pe un sheet nou
        self.generate_histogram(df_tables, workbook)

        # Adăugare Status Analysis
        self.generate_status_analysis(df_tables, workbook)

         # Adăugare Cycle Analysis
        self.generate_cycle_histograms(df_tables, workbook)

        self.generate_table_performance_report(df_tables, workbook)

        # Generarea histogramei pentru oameni
        self.generate_people_histogram(df_people, workbook)
        # Salvarea fișierului Excel
        workbook.save(self.excel_file)

    def generate_people_histogram(self, df_people, workbook):
        """Generează histograma pentru numărul de oameni și o adaugă în Excel."""
        # Gruparea numărului de oameni pe intervale orare
        df_people['time_slot'] = pd.to_datetime(df_people['time']).dt.floor('h')  # Folosește 'h' în loc de 'H'
        
        # Calcularea numărului maxim de oameni pe intervale orare
        people_max_count = df_people.groupby('time_slot')['people_count'].max().reset_index()

        # Crearea histogramei
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.bar(people_max_count['time_slot'].astype(str), people_max_count['people_count'], color='skyblue')
        ax.set_xlabel('Time (Hourly intervals)')
        ax.set_ylabel('Max Number of People')
        ax.set_title('Max Number of People per Hour')

        # Salvăm graficul temporar
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmpfile:
            plt.savefig(tmpfile.name, format='png')
            plt.close()

            # Adăugăm graficul în sheet-ul People Statistics
            ws = workbook.create_sheet('People Statistics')
            img = Image(tmpfile.name)
            ws.add_image(img, 'A1')

        # Scrierea numărului maxim de oameni în intervalele orare în Excel
        sheet = workbook.create_sheet('Max People per Hour')
        sheet.append(['Time', 'Max People Count'])  # Adăugăm antetul

        # Adăugăm fiecare interval orar și numărul maxim de oameni
        for index, row in people_max_count.iterrows():
            sheet.append([row['time_slot'].strftime('%H:%M'), row['people_count']])
            self.people_number_list.append(row['people_count'])

        print("Max number of people per hour added to Excel.")



    def generate_histogram(self, df, workbook):
        """Generează histograma și o adaugă în Excel."""
        statuses = {'need to clean': 'red', 'available': 'green', 'ready to order': 'orange', 'eating': 'blue'}
        table_durations = {table_id: {status: 0 for status in statuses} for table_id in df['table_id'].unique()}

        # Calcularea duratelor totale pentru fiecare status al fiecărei mese
        for _, row in df.iterrows():
            table_id = row['table_id']
            status = row['status']
            duration = self.convert_duration(row['duration'])
            table_durations[table_id][status] += duration

        # Pregătirea datelor pentru histogramă
        table_ids = sorted(table_durations.keys())
        status_names = list(statuses.keys())
        bar_width = 0.2
        index = range(len(table_ids))

        # Crearea graficului
        fig, ax = plt.subplots(figsize=(10, 6))
        for i, status in enumerate(status_names):
            durations = [table_durations[table_id][status] for table_id in table_ids]
            ax.bar([x + i * bar_width for x in index], durations, bar_width, color=statuses[status], label=status)
        # Configurarea axei y
        max_duration = max(durations)
        yticks = np.linspace(0, max_duration, num=10)  # Creează 5 tick-uri distribuite uniform
        formatted_yticks = [self.format_duration(int(tick)) for tick in yticks]

        ax.set_xlabel('Table ID')
        ax.set_ylabel('Duration (seconds)')
        ax.set_title('Table Status Duration Histogram')
        ax.set_xticks([x + bar_width * 1.5 for x in index])
        ax.set_xticklabels([f"Table {table_id}" for table_id in table_ids], rotation=45)
        ax.set_yticks(yticks)
        ax.set_yticklabels(formatted_yticks)
        ax.legend()
        plt.tight_layout()

        # Salvează graficul temporar
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmpfile:
            plt.savefig(tmpfile.name, format='png')
            plt.close()

            # Adaugă graficul în Excel
            ws = workbook.create_sheet('Histogram')
            img = Image(tmpfile.name)
            ws.add_image(img, 'A1')

    def generate_average_histogram(self, avg_durations, workbook):
        """Generează o histogramă pe baza mediilor și un grafic procentual, apoi le adaugă în același sheet din Excel."""
        statuses = {'need to clean': 'red', 'available': 'green', 'ready to order': 'orange', 'eating': 'blue'}

        # Pregătim datele pentru histogramă
        labels = []
        values = []
        colors = []
        for status, avg in avg_durations.items():
            if avg > 0:
                labels.append(status)
                values.append(avg)
                colors.append(statuses[status])

        # Creăm histograma
        fig, ax = plt.subplots(figsize=(8, 5))
        ax.bar(labels, values, color=colors)
        ax.set_title('Average Durations by Status')
        ax.set_ylabel('Duration (seconds)')
        ax.set_xlabel('Status')

        # Salvăm histograma temporar
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmpfile_hist:
            plt.savefig(tmpfile_hist.name, format='png')
            plt.close()

            # Adăugăm histograma în sheet
            ws = workbook.create_sheet('Average Analysis')
            img_hist = Image(tmpfile_hist.name)
            ws.add_image(img_hist, 'A1')  # Poziționăm histograma în stânga

        # Creăm graficul procentual
        fig, ax = plt.subplots(figsize=(6, 6))
        ax.pie(values, labels=labels, colors=colors, autopct='%1.1f%%', startangle=140)
        ax.set_title('Percentage of Average Durations by Status')

        # Salvăm graficul procentual temporar
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmpfile_pie:
            plt.savefig(tmpfile_pie.name, format='png')
            plt.close()

            # Adăugăm graficul pie chart în dreapta histogramei
            img_pie = Image(tmpfile_pie.name)
            ws.add_image(img_pie, 'P1')  # Poziționăm imaginea în coloana K, pe același rând

    def generate_status_analysis(self, df, workbook):
        """
        Generează un sheet cu:
        1. Timpul maxim și minim pentru fiecare status.
        2. O diagramă de tip box plot care arată distribuția timpilor pentru fiecare status.
        """
        import seaborn as sns

        # Creăm un sheet nou
        ws = workbook.create_sheet('Status Analysis')

        # 1. Calculăm timpii maximi și minimi pentru fiecare status
        statuses = df['status'].unique()
        ws.append(['Status', 'Max Duration', 'Min Duration'])  # Header
        for status in statuses:
            # Convertim durata în secunde pentru calcul
            status_data = df[df['status'] == status]
            if not status_data.empty:
                max_duration = status_data['duration'].apply(self.convert_duration).max()
                min_duration = status_data['duration'].apply(self.convert_duration).min()
                ws.append([status, self.format_duration(max_duration), self.format_duration(min_duration)])
            else:
                ws.append([status, "N/A", "N/A"])

        # 2. Generăm un box plot pentru distribuția timpilor
        fig, ax = plt.subplots(figsize=(10, 6))
        df['duration_seconds'] = df['duration'].apply(self.convert_duration)  # Adăugăm o coloană temporară
        sns.boxplot(x='status', y='duration_seconds', data=df, hue='status', palette='Set2', ax=ax, legend=False)

        ax.set_title('Duration Distribution by Status')
        ax.set_ylabel('Duration (seconds)')
        ax.set_xlabel('Status')

        # Salvăm graficul box plot temporar
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmpfile_boxplot:
            plt.savefig(tmpfile_boxplot.name, format='png')
            plt.close()

            # Adăugăm graficul în sheet
            img_boxplot = Image(tmpfile_boxplot.name)
            ws.add_image(img_boxplot, 'E1')  # Poziționăm box plot-ul în sheet

    def calculate_avg_durations(self, df):
        """Calculează media duratelor pentru fiecare status, pe baza datelor din DataFrame."""
        statuses = ['need to clean', 'available', 'ready to order', 'eating']
        avg_durations = {}

        # Calculăm suma și numărul de intrări pentru fiecare status
        for status in statuses:
            filtered_data = df[df['status'] == status]
            if not filtered_data.empty:
                total_duration = filtered_data['duration'].apply(self.convert_duration).sum()
                avg_duration = total_duration / len(filtered_data)
                avg_durations[status] = avg_duration
            else:
                avg_durations[status] = 0  # Dacă nu sunt date pentru acest status, media este 0

        return avg_durations
        
    def add_time_slot(self, df):
        """Adaugă intervalul orar pentru fiecare intrare pe baza duratei."""
        from datetime import datetime

        def get_time_slot(row):
            """Extrage intervalul orar pe baza orei de început."""
            # Convertește start_time în datetime
            start_time = datetime.strptime(row['start_time'], '%Y-%m-%d %H:%M:%S')
            hour = start_time.hour
            if hour < 6:
                return 'Night'
            elif hour < 12:
                return 'Morning'
            elif hour < 18:
                return 'Afternoon'
            else:
                return 'Evening'

        df['time_slot'] = df.apply(get_time_slot, axis=1)
        return df
    
    def generate_cycle_histograms(self, df, workbook):
        """Generează histogramele pentru durata ciclurilor fiecărei mese și durata medie a ciclurilor."""
        import matplotlib.pyplot as plt
        from matplotlib.colors import TABLEAU_COLORS
        import tempfile
        from openpyxl.drawing.image import Image
        
        # 1. Calcularea duratelor ciclurilor
        total_durations, individual_durations = self.calculate_table_cycle_duration(df)
        
        # Pregătirea unei palete de culori pentru prima histogramă
        colors = list(TABLEAU_COLORS.values())
        
        # Generăm histograma pentru fiecare masă cu duratele ciclurilor individuale
        fig, ax = plt.subplots(figsize=(12, 8))
        
        for i, (table_id, durations) in enumerate(individual_durations.items()):
            color_index = 0
            for j, duration in enumerate(durations):
                ax.bar(f"{table_id} - Cycle {j+1}", duration, color=colors[color_index % len(colors)])
                color_index += 1
        
        ax.set_xlabel('Tables and Cycles')
        ax.set_ylabel('Cycle Duration (seconds)')
        ax.set_title('Cycle Duration for Each Table and Cycle')
        plt.xticks(rotation=45, ha='right')
        
        # Salvarea primei histograme
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmpfile_cycle:
            plt.savefig(tmpfile_cycle.name, format='png')
            plt.close()
            
            if 'cycle_duration_histograms' not in workbook.sheetnames:
                ws = workbook.create_sheet('cycle_duration_histograms')
            else:
                ws = workbook['Cycle Duration Histograms']
            
            img_cycle = Image(tmpfile_cycle.name)
            ws.add_image(img_cycle, 'A1')
        
        # Calcularea mediei ciclurilor pentru fiecare masă
        avg_durations = {table_id: sum(durations) / len(durations) if durations else 0
                        for table_id, durations in individual_durations.items()}
        
        # Generăm histograma cu duratele medii
        fig, ax = plt.subplots(figsize=(12, 8))
        ax.bar(avg_durations.keys(), avg_durations.values(), color='green')
        ax.set_xlabel('Table ID')
        ax.set_ylabel('Average Cycle Duration (seconds)')
        ax.set_title('Average Cycle Duration for Each Table')
        plt.xticks(rotation=45, ha='right')
        
        # Salvarea celei de-a doua histograme
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmpfile_avg:
            plt.savefig(tmpfile_avg.name, format='png')
            plt.close()
            
            img_avg = Image(tmpfile_avg.name)
            ws.add_image(img_avg, 'A20')  # Păstrăm o distanță între imagini

    def calculate_table_cycle_duration(self, df):
        """Calculază durata totală a ciclurilor fiecărei mese și returnează și lista duratelor individuale ale ciclurilor."""
        table_cycle_durations = {}  # Durata totală a ciclurilor pentru fiecare masă
        all_cycle_durations = {}  # Duratele individuale ale ciclurilor pentru fiecare masă
        
        for table_id in df['table_id'].unique():
            table_data = df[df['table_id'] == table_id]
            available_to_clean_duration = 0
            cycle_durations = []  # Lista pentru duratele ciclurilor pentru fiecare masă
            is_in_cycle = False
            
            for _, row in table_data.iterrows():
                if row['status'] == 'available':
                    is_in_cycle = True
                    available_to_clean_duration = 0  # Resetează durata ciclului pentru un nou ciclu
                if is_in_cycle:
                    available_to_clean_duration += self.convert_duration(row['duration'])
                if row['status'] == 'need to clean' and is_in_cycle:
                    cycle_durations.append(available_to_clean_duration)
                    is_in_cycle = False
            
            table_cycle_durations[table_id] = sum(cycle_durations)  # Suma tuturor ciclurilor pentru masă
            all_cycle_durations[table_id] = cycle_durations  # Duratele individuale
            
        return table_cycle_durations, all_cycle_durations

    def generate_table_performance_report(self, df, workbook):
        """Generează raportul de performanță al meselor și îl salvează într-un sheet Excel."""
        import pandas as pd
        from openpyxl import Workbook
        
        # Ponderi pentru calculul scorului de performanță
        weights = [0.4, 0.3, 0.2, 0.1]

        # Pregătim tabelul rezultat
        rows = []
        unique_table_ids = df['table_id'].unique()

        for table_id in unique_table_ids:
            table_data = df[df['table_id'] == table_id]
            
            # Calculăm statistici pentru fiecare masă
            total_time = table_data['duration'].apply(self.convert_duration).sum()
            total_cycles = len(table_data[table_data['status'] == 'need to clean'])
            time_eating = table_data[table_data['status'] == 'eating']['duration'].apply(self.convert_duration).sum()
            time_ready_to_order = table_data[table_data['status'] == 'ready to order']['duration'].apply(self.convert_duration).sum()
            time_need_to_clean = table_data[table_data['status'] == 'need to clean']['duration'].apply(self.convert_duration).sum()
            time_available = table_data[table_data['status'] == 'available']['duration'].apply(self.convert_duration).sum()
        

            idle_time_ratio = time_available / total_time if total_time else 0
            cleaning_ratio = time_need_to_clean / total_time if total_time else 0
            cycles_per_hour = total_cycles / (total_time / 3600) if total_time else 0
            utilization_rate = (time_eating + time_ready_to_order) / total_time if total_time else 0
            
            # Calculăm scorul de performanță
            performance_score = (
                (weights[0] * cycles_per_hour) +
                (weights[1] * utilization_rate) -
                (weights[2] * cleaning_ratio) -
                (weights[3] * idle_time_ratio)
            )

            self.tables_performance_list.append(performance_score)      
            self.idle_time_ratio_list.append(idle_time_ratio)
            self.cleaning_ratio_list.append(cleaning_ratio)
            self.cycles_per_hour_list.append(cycles_per_hour)
            self.utilization_rate_list.append(utilization_rate)

            total_time = self.format_duration(total_time)
            total_cycles = self.format_duration(total_cycles)
            time_eating = self.format_duration(time_eating)
            time_ready_to_order = self.format_duration(time_ready_to_order)
            time_need_to_clean = self.format_duration(time_need_to_clean)
            time_available = self.format_duration(time_available)
            # Adăugăm datele în rânduri
            rows.append({
                'Table ID': table_id,
                'Total Time': total_time,
                'Total Cycles': total_cycles,
                'Time Eating': time_eating,
                'Time Ready to Order': time_ready_to_order,
                'Time Need to Clean': time_need_to_clean,
                'Time Available': time_available,
                'Idle Time Ratio': idle_time_ratio,
                'Cleaning Ratio': cleaning_ratio,
                'Cycles per Hour': cycles_per_hour,
                'Utilization Rate': utilization_rate,
                'Performance Score': performance_score
            })

        # Creăm un DataFrame pentru rezultatul final
        result_df = pd.DataFrame(rows)

        # Scriem rezultatul într-un nou sheet al workbook-ului
        sheet_name = 'Table Performance'
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

        worksheet = workbook.create_sheet(sheet_name)

        # Scrierea header-ului
        for col_index, col_name in enumerate(result_df.columns, start=1):
            worksheet.cell(row=1, column=col_index, value=col_name)

        # Scrierea datelor
        for row_index, row in enumerate(result_df.itertuples(index=False), start=2):
            for col_index, value in enumerate(row, start=1):
                worksheet.cell(row=row_index, column=col_index, value=value)

        # Adăugăm filtre pentru coloane
        start_column = 2  # Coloanele de la 2 încolo vor avea filtre
        end_column = len(result_df.columns)  # Ultima coloană a tabelului
        worksheet.auto_filter.ref = worksheet.dimensions  # Adaugă filtre pentru întregul tabel

    def calculate_averages(self):
        """
        Calculează media pentru fiecare din listele de performanță.
        Media pentru people_number este un număr întreg, iar celelalte sunt floaturi cu maxim 4 zecimale.
        Returnează un dicționar cu mediile pentru fiecare listă.
        """
        # Dicționar pentru a stoca rezultatele mediilor
        averages = {}

        # Lista de tuple (numele listei, lista corespunzătoare)
        lists = [
            ("tables_performance", self.tables_performance_list),
            ("people_number", self.people_number_list),
            ("idle_time_ratio", self.idle_time_ratio_list),
            ("cleaning_ratio", self.cleaning_ratio_list),
            ("cycles_per_hour", self.cycles_per_hour_list),
            ("utilization_rate", self.utilization_rate_list)
        ]

        for list_name, data_list in lists:
            if data_list:  # Verificăm dacă lista nu este goală
                average = sum(data_list) / len(data_list)
                if list_name == "people_number":
                    averages[list_name] = int(round(average))  # Convertim media în întreg
                else:
                    averages[list_name] = round(average, 4)  # Rotunjim la 4 zecimale
            else:
                averages[list_name] = None  # Dacă lista este goală, returnăm None

        return averages
    
    def save_average_statistics(self):
        """
        Generează un fișier JSON cu mediile calculate pentru listele de statistici,
        utilizând data curentă ca cheie unică în dicționar.
        Creează folderul `data/outputs/daily_report/<%Y-%m-%d>` dacă nu există și
        salvează fișierul `average_statistics.json` în acest folder.
        """
        # Calculăm mediile folosind funcția calculate_averages
        averages = self.calculate_averages()

        # Obține data curentă în formatul YYYY-MM-DD
        current_date = datetime.now().strftime("%Y-%m-%d")

        # Creează un dicționar cu data curentă ca cheie unică
        data_to_save = {current_date: averages}

        # Creează calea completă a folderului
        output_dir = os.path.join("data", "outputs", "daily_report", current_date)
        os.makedirs(output_dir, exist_ok=True)  # Creează folderul dacă nu există

        # Creează calea completă pentru fișierul JSON
        json_file_path = os.path.join(output_dir, "average_statistics.json")

        # Salvează dicționarul în fișierul JSON
        with open(json_file_path, "w") as json_file:
            json.dump(data_to_save, json_file, indent=4)

        print(f"Fișierul JSON {json_file_path} a fost generat cu succes.")

    @staticmethod
    def convert_duration(duration_str):
        """Convertește durata în secunde."""
        hours, minutes, seconds = map(int, duration_str.split(':'))
        return hours * 3600 + minutes * 60 + seconds
    
    @staticmethod
    def format_duration(seconds):
        """Formatează durata în hh:mm:ss."""
        try:
            seconds = int(seconds)  # Conversie în număr întreg
        except ValueError:
            raise ValueError(f"Durata trebuie să fie un număr întreg, dar s-a primit: {seconds}")
        
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        seconds = seconds % 60
        return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"

        